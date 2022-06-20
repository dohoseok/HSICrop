import os
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import confusion_matrix
from openpyxl import Workbook
from PIL import Image

def save_plt_fig(x_data, y_data, x_label, y_label, save_path):
    plt.clf()
    plt.plot(x_data, y_data)
    plt.grid()
    plt.xlabel(x_label)
    plt.ylabel(y_label)

    foldername = os.path.basename(os.path.normpath(save_path))
    filename = os.path.join(save_path, f'{foldername}_{y_label}.png')
    plt.savefig(filename)


def metrics(prediction, target, ignored_labels=[], n_classes=None):
    """Compute and print metrics (accuracy, confusion matrix and F1 scores).

    Args:
        prediction: list of predicted labels
        target: list of target labels
        ignored_labels (optional): list of labels to ignore, e.g. 0 for undef
        n_classes (optional): number of classes, max(target) by default
    Returns:
        accuracy, F1 score by class, confusion matrix
    """
    ignored_mask = np.zeros(target.shape[:2], dtype=np.bool)
    for l in ignored_labels:
        ignored_mask[target == l] = True
    ignored_mask = ~ignored_mask
    target = target[ignored_mask]
    prediction = prediction[ignored_mask]

    results = {}

    n_classes = np.max(target) + 1 if n_classes is None else n_classes

    cm = confusion_matrix(
        target,
        prediction,
        labels=range(n_classes))

    results["Confusion matrix"] = cm

    # Compute global accuracy
    total = np.sum(cm)
    accuracy = sum([cm[x][x] for x in range(len(cm))])
    accuracy *= 100 / float(total)

    results["Accuracy"] = accuracy

    # Compute accuracy for each class
    accuracies = np.zeros(len(cm))
    ious = np.zeros(len(cm))
    for i in range(len(cm)):
        acc = cm[i,i] / np.sum(cm[i,:])
        iou = cm[i,i] / (np.sum(cm[i,:]) + np.sum(cm[:,i]) - cm[i,i])

        accuracies[i] = acc
        ious[i] = iou

    results["accuracies"] = accuracies
    results["ious"] = ious

    # Compute F1 score
    F1scores = np.zeros(len(cm))
    for i in range(len(cm)):
        try:
            F1 = 2. * cm[i, i] / (np.sum(cm[i, :]) + np.sum(cm[:, i]))
        except ZeroDivisionError:
            F1 = 0.
        F1scores[i] = F1

    results["F1 scores"] = F1scores

    # Compute kappa coefficient
    pa = np.trace(cm) / float(total)
    pe = np.sum(np.sum(cm, axis=0) * np.sum(cm, axis=1)) / \
        float(total * total)
    kappa = (pa - pe) / (1 - pe)
    results["Kappa"] = kappa

    return results


def save_results(results, result_filename): #vis, label_values=None, agregated=False):
    cm = results["Confusion matrix"]
    accuracy = results["Accuracy"]
    accuracies = results["accuracies"]
    ious = results["ious"]

    AA = np.mean(accuracies)
    mIou = np.mean(ious)
    fIou = np.sum(np.multiply(ious, np.sum(cm, axis=1))) / np.sum(cm)

    # print("iou: ", ious)
    # print("Acc: ", accuracies)
    print("result_filename", result_filename)
    result_str = f'{mIou:.4f},{fIou:.4f},{accuracy/100.:.4f},{AA:.4f}'
    for iou in ious:
        result_str += f',{iou:.4f}'
    for acc in accuracies:
        result_str += f',{acc:.4f}'
    print('Result All :' , result_str)
    # print(f'Result All : {mIou}, {fIou}, {accuracy/100.}, {AA}, {str(ious)[1:-1]}, {str(accuracies)[1:-1]}')

    NUM_CLASSES = len(ious)
    wb = Workbook()
    ws = wb.active

    for col in range(NUM_CLASSES):
        for row in range(NUM_CLASSES):
            ws.cell(row=row + 1, column=col + 1, value=cm[row][col])

    for class_idx in range(NUM_CLASSES):
        ws.cell(row=NUM_CLASSES + 1, column=class_idx + 1, value=accuracies[class_idx])
        ws.cell(row=NUM_CLASSES + 2, column=class_idx + 1, value=ious[class_idx])

    ws.cell(row=NUM_CLASSES + 3, column=1, value="mIou")
    ws.cell(row=NUM_CLASSES + 4, column=1, value="fIou")
    ws.cell(row=NUM_CLASSES + 5, column=1, value="Average Acc")
    ws.cell(row=NUM_CLASSES + 6, column=1, value="Overall Acc")
    ws.cell(row=NUM_CLASSES + 3, column=2, value=mIou)
    ws.cell(row=NUM_CLASSES + 4, column=2, value=fIou)
    ws.cell(row=NUM_CLASSES + 5, column=2, value=AA)
    ws.cell(row=NUM_CLASSES + 6, column=2, value=accuracy/100)

    wb.save(result_filename)


def save_visible(predictions, save_path, test_txt):
    os.makedirs(save_path, exist_ok=True)

    pred_vis = predictions.reshape(-1, 128, 128)
    print(pred_vis.shape)

    palette = [
        0, 0, 0,
        255	,	0	,	0	,	#1
        0	,	255	,	0	,	#2
        0	,	0	,	255	,	#3
        255	,	255	,	0	,	#4
        255	,	0	,	255	,	#5
        0	,	255	,	255	,	#6
        255	,	242	,	204	,	#7
        208	,	166	,	82	,	#8
        128	,	128	,	128	,	#9
        217	,	155	,	210	,	#10
        153	,	153	,	255	,	#11
        159	,	242	,	255	,	#12
    ]

    lines = open(test_txt, 'rt').readlines()
    for line, pred in zip(lines, pred_vis):
        filename = line.strip()
        pred = np.uint8(pred)
        # print(pred)
        image = Image.fromarray(pred)
        image.putpalette(palette)
        image.save(os.path.join(save_path, filename + "_vis.png"))